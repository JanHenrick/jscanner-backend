from openpyxl import Workbook
import os

def convert_text_to_excel(text: str, filename: str) -> str:
    output_path = f"outputs/{filename}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Converted Data"
    
    # Header
    ws['A1'] = "PHDCIScanner - Converted Document"
    ws['A1'].font = ws['A1'].font.copy(bold=True)
    
    # Content - each line = one row
    row = 3
    for line in text.split('\n'):
        if line.strip():
            ws.cell(row=row, column=1, value=line)
            row += 1
    
    # Auto-fit column width
    ws.column_dimensions['A'].width = 80
    
    wb.save(output_path)
    return output_path